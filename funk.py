def parsing(adresa_komirku):
    from openpyxl import load_workbook, Workbook

    book = load_workbook(filename="15.05.2023.xlsx")
    sheet = book.active

    wb = Workbook()
    shop_name = f'{sheet[adresa_komirku + "1"].value}.xlsx'
    wb.save(shop_name)

    book_shop = load_workbook(shop_name)
    ws = book_shop["Sheet"]
    for i in range(3, sheet.max_row):
        if not sheet[adresa_komirku + str(i)].value:
            continue
        ws["A" + str(i)] = sheet["A" + str(i)].value
        ws["B" + str(i)] = sheet[adresa_komirku + str(i)].value
        ws["C" + str(i)] = sheet["Y" + str(i)].value

    book_shop.save(shop_name)
    book_shop.close()
